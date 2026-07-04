"""Autonomous municipal debt collection agent.

Version 1 is intentionally dependency-light and file based so it can run in a
cron job, application worker, or manually from the command line. It persists run
state, audit logs, imported records, recommendations, collector queues, and the
manager briefing as JSON artifacts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import time
import uuid
from contextlib import contextmanager
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from municipal_ai_os.field_activity_attribution import FieldActivityAttributionEngine
from municipal_ai_os.financial_intelligence import FinancialIntelligenceDashboard


WORKFLOW_STEPS = [
    "detect_files",
    "import_data",
    "clean_data",
    "validate_records",
    "detect_duplicates",
    "apply_rules",
    "run_ai_assistant",
    "generate_recommendations",
    "generate_work_queue",
    "generate_daily_snapshot",
    "generate_manager_briefing",
    "generate_executive_intelligence",
    "generate_field_activity_attribution",
    "generate_financial_intelligence",
]

REGION_1_SCHEMA_ALIASES = {
    "debtor_id": ["شماره_پرونده", "کد_شناسایی"],
    "debtor": ["نام_متصدی"],
    "address": ["نشانی_واحد_صنفی"],
    "mobile": ["شماره_تماس"],
    "debt_amount": ["بدهی_معوقه"],
    "due_date": ["تاریخ_پرداخت"],
    "business_category": ["شغل_واحد"],
    "legal_status": ["وضعیت_حقوقی", "وضعیت_پرونده"],
    "zone": ["منطقه", "ناحیه"],
}

MUNICIPAL_ACTION_QUEUES = {
    "CALL": "daily_call_queue.json",
    "FIELD_VISIT": "daily_visit_queue.json",
    "NOTICE": "daily_notice_queue.json",
    "LEGAL": "legal_queue.json",
    "FOLLOW_UP": "daily_follow_up_queue.json",
    "IGNORE": "ignore_queue.json",
}

DEFAULT_SNAPSHOT_COMPARISON_RULES = {
    "paid_debt_threshold": 0,
    "minimum_debt_change": 0,
    "cleared_legal_statuses": ["paid", "cleared", "closed", "resolved"],
}

DEFAULT_COLLECTOR_PERFORMANCE_RULES = {
    "unchanged_many_threshold": 10,
    "unchanged_ratio_alert": 0.7,
    "repeated_visit_no_progress_threshold": 2,
    "notice_no_payment_threshold": 5,
    "high_priority_ignored_score": 80,
    "score_weights": {
        "base": 70,
        "improved": 8,
        "paid": 10,
        "worsened": -12,
        "unchanged": -2,
        "overdue": -5,
    },
}


@dataclass(frozen=True)
class DebtRecord:
    record_id: str
    debtor_id: str
    debtor: str
    address: str
    mobile: str | None
    region: str
    zone: str
    collector: str
    debt_amount: float
    business_category: str | None = None
    legal_status: str | None = None
    due_date: str | None = None
    last_contact_date: str | None = None
    status: str = "open"
    source_file: str = ""
    row_number: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class Recommendation:
    record_id: str
    debtor: str
    address: str
    region: str
    collector: str
    debt_amount: float
    recommended_action: str
    priority: str
    explanation: str
    estimated_success_probability: float


@dataclass
class RunState:
    run_id: str
    mode: str
    start_time: str
    end_time: str | None = None
    duration_seconds: float | None = None
    processed_files: list[str] = field(default_factory=list)
    processed_debt_records: int = 0
    recommendations_generated: int = 0
    executive_recommendations_generated: int = 0
    warnings: list[str] = field(default_factory=list)
    failures: list[str] = field(default_factory=list)
    completed_steps: list[str] = field(default_factory=list)
    last_successful_step: str | None = None
    status: str = "running"


class DuplicateRunError(RuntimeError):
    """Raised when another collection agent run is already active."""


class CollectionAgent:
    """Runs the autonomous collection workflow with recovery and audit logging."""

    def __init__(self, workspace: str | Path = "var/collection_agent") -> None:
        self.workspace = Path(workspace)
        self.inbox_dir = self.workspace / "inbox"
        self.archive_dir = self.workspace / "archive"
        self.runs_dir = self.workspace / "runs"
        self.state_dir = self.workspace / "state"
        self.audit_dir = self.workspace / "audit"
        self.lock_path = self.state_dir / "agent.lock"
        for directory in [
            self.inbox_dir,
            self.archive_dir,
            self.runs_dir,
            self.state_dir,
            self.audit_dir,
        ]:
            directory.mkdir(parents=True, exist_ok=True)

    @contextmanager
    def _run_lock(self) -> Iterable[None]:
        try:
            fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode())
            os.close(fd)
        except FileExistsError as exc:
            raise DuplicateRunError(
                f"Collection agent is already running: {self.lock_path}"
            ) from exc
        try:
            yield
        finally:
            self.lock_path.unlink(missing_ok=True)

    def run(self, mode: str = "manual", resume_run_id: str | None = None) -> RunState:
        """Execute or resume the daily collection workflow."""
        with self._run_lock():
            state = self._load_or_create_state(mode, resume_run_id)
            artifacts: dict[str, Any] = {}
            started = datetime.fromisoformat(state.start_time)
            try:
                for step in WORKFLOW_STEPS:
                    if step in state.completed_steps:
                        artifacts.update(self._load_step_artifacts(state.run_id, step))
                        self._audit(state.run_id, "step_skipped", {"step": step})
                        continue
                    self._audit(state.run_id, "step_started", {"step": step})
                    artifacts.update(getattr(self, f"_{step}")(state, artifacts))
                    state.completed_steps.append(step)
                    state.last_successful_step = step
                    self._save_step_artifacts(state.run_id, step, artifacts)
                    self._save_state(state)
                    self._audit(state.run_id, "step_completed", {"step": step})
                state.status = "completed"
            except Exception as exc:  # noqa: BLE001 - failures must be captured for recovery
                state.status = "failed"
                state.failures.append(f"{type(exc).__name__}: {exc}")
                self._audit(state.run_id, "run_failed", {"error": state.failures[-1]})
                raise
            finally:
                ended = datetime.now(timezone.utc)
                state.end_time = ended.isoformat()
                state.duration_seconds = round((ended - started).total_seconds(), 3)
                self._save_state(state)
                self._audit(state.run_id, "run_finished", asdict(state))
            return state

    def _load_or_create_state(self, mode: str, resume_run_id: str | None) -> RunState:
        if resume_run_id:
            state_file = self.runs_dir / resume_run_id / "run_state.json"
            if not state_file.exists():
                raise FileNotFoundError(f"Run state not found for {resume_run_id}")
            return RunState(**json.loads(state_file.read_text()))
        run_id = f"{date.today().isoformat()}-{uuid.uuid4().hex[:8]}"
        state = RunState(
            run_id=run_id, mode=mode, start_time=datetime.now(timezone.utc).isoformat()
        )
        self._save_state(state)
        self._audit(run_id, "run_started", asdict(state))
        return state

    def _detect_files(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        seen = self._seen_file_hashes()
        files = []
        for path in sorted(self.inbox_dir.glob("*")):
            if path.suffix.lower() not in {".xlsx", ".csv"} or not path.is_file():
                continue
            digest = self._sha256(path)
            if digest not in seen:
                files.append({"path": str(path), "name": path.name, "sha256": digest})
        state.processed_files = [item["name"] for item in files]
        return {"files": files}

    def _import_data(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        raw_rows: list[dict[str, Any]] = []
        for file_info in artifacts.get("files", []):
            path = Path(file_info["path"])
            rows = self._read_xlsx(path) if path.suffix.lower() == ".xlsx" else self._read_csv(path)
            for row_number, row in enumerate(rows, start=2):
                row["source_file"] = path.name
                row["row_number"] = row_number
                raw_rows.append(row)
        return {"raw_rows": raw_rows}

    def _clean_data(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        cleaned = []
        for row in artifacts.get("raw_rows", []):
            normalized = {self._normalize_key(k): self._clean_value(v) for k, v in row.items()}
            self._apply_schema_mapping(normalized)
            normalized["debt_amount"] = self._money(normalized.get("debt_amount", 0))
            cleaned.append(normalized)
        return {"cleaned_rows": cleaned}

    def _validate_records(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        records: list[dict[str, Any]] = []
        required = ["debtor", "address", "debt_amount"]
        for row in artifacts.get("cleaned_rows", []):
            warnings = [f"missing {field}" for field in required if not row.get(field)]
            if row.get("debt_amount", 0) <= 0:
                warnings.append("debt_amount must be positive")
            if warnings:
                state.warnings.append(
                    f"{row.get('source_file')}:{row.get('row_number')} {', '.join(warnings)}"
                )
                continue
            record = DebtRecord(
                record_id=self._record_id(row),
                debtor_id=self._debtor_id(row),
                debtor=str(row["debtor"]),
                address=str(row["address"]),
                mobile=str(row["mobile"]) if row.get("mobile") else None,
                region=str(row.get("region") or "Unassigned"),
                zone=str(row.get("zone") or row.get("region") or "Unassigned"),
                collector=str(row.get("collector") or "Unassigned"),
                debt_amount=float(row["debt_amount"]),
                business_category=str(row["business_category"])
                if row.get("business_category")
                else None,
                legal_status=str(row["legal_status"]) if row.get("legal_status") else None,
                due_date=row.get("due_date"),
                last_contact_date=row.get("last_contact_date"),
                status=str(row.get("status") or "open"),
                source_file=str(row.get("source_file") or ""),
                row_number=int(row.get("row_number") or 0),
                warnings=warnings,
            )
            records.append(asdict(record))
        state.processed_debt_records = len(records)
        return {"records": records}

    def _detect_duplicates(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        unique, duplicates, seen = [], [], set()
        for record in artifacts.get("records", []):
            key = (record["debtor"].lower(), record["address"].lower(), record["debt_amount"])
            if key in seen:
                duplicates.append(record)
                state.warnings.append(f"duplicate record skipped: {record['record_id']}")
            else:
                seen.add(key)
                unique.append(record)
        return {"records": unique, "duplicates": duplicates}

    def _apply_rules(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        scored = []
        today = date.today()
        for record in artifacts.get("records", []):
            amount = record["debt_amount"]
            days_overdue = self._days_since(record.get("due_date"), today)
            no_contact_days = self._days_since(record.get("last_contact_date"), today)
            score = min(
                100, int(amount / 100) + max(days_overdue, 0) + max(no_contact_days // 3, 0)
            )
            priority = "high" if score >= 80 else "medium" if score >= 40 else "low"
            action = (
                "field visit"
                if priority == "high"
                else "phone call"
                if priority == "medium"
                else "courtesy reminder"
            )
            scored.append({**record, "score": score, "priority": priority, "rule_action": action})
        return {"scored_records": scored}

    def _run_ai_assistant(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        assisted = []
        for record in artifacts.get("scored_records", []):
            probability = max(
                0.1,
                min(
                    0.95,
                    0.72 - (record["score"] / 300) + (0.08 if record["debt_amount"] < 500 else 0),
                ),
            )
            explanation = (
                f"{record['priority'].title()} priority based on amount ${record['debt_amount']:.2f}, "
                f"overdue age, and recent contact history. Recommended {record['rule_action']}."
            )
            assisted.append(
                {**record, "success_probability": round(probability, 2), "explanation": explanation}
            )
        return {"assisted_records": assisted}

    def _generate_recommendations(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        recs = []
        for record in artifacts.get("assisted_records", []):
            municipal_action = self._municipal_action(record)
            recommendation = asdict(
                Recommendation(
                    record_id=record["record_id"],
                    debtor=record["debtor"],
                    address=record["address"],
                    region=record["region"],
                    collector=record["collector"],
                    debt_amount=record["debt_amount"],
                    recommended_action=municipal_action,
                    priority=record["priority"],
                    explanation=record["explanation"],
                    estimated_success_probability=record["success_probability"],
                )
            )
            recommendation["priority_score"] = record["score"]
            recommendation["zone"] = self._district_from_record(record)
            recs.append(recommendation)
        state.recommendations_generated = len(recs)
        return {"recommendations": recs}

    def _generate_work_queue(self, state: RunState, artifacts: dict[str, Any]) -> dict[str, Any]:
        queues: dict[str, list[dict[str, Any]]] = {action: [] for action in MUNICIPAL_ACTION_QUEUES}
        rank = {"high": 0, "medium": 1, "low": 2}
        for rec in sorted(
            artifacts.get("recommendations", []),
            key=lambda r: (rank[r["priority"]], -r["debt_amount"]),
        ):
            queues[rec["recommended_action"]].append(rec)
        for action, filename in MUNICIPAL_ACTION_QUEUES.items():
            self._write_json(state.run_id, filename, queues[action])
        self._write_json(state.run_id, "collector_work_queues.json", queues)
        dashboard = self._manager_assignment_dashboard(queues)
        self._write_json(state.run_id, "manager_dashboard.json", dashboard)
        return {"work_queues": queues, "manager_dashboard": dashboard}

    def _generate_daily_snapshot(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        rules = self._load_json_config(
            "snapshot_comparison_rules.json", DEFAULT_SNAPSHOT_COMPARISON_RULES
        )
        performance_rules = self._load_json_config(
            "collector_performance_rules.json", DEFAULT_COLLECTOR_PERFORMANCE_RULES
        )
        recommendations = {rec["record_id"]: rec for rec in artifacts.get("recommendations", [])}
        snapshot = [
            self._snapshot_record(record, recommendations.get(record["record_id"]), state.run_id)
            for record in artifacts.get("scored_records", [])
        ]
        previous_snapshot, previous_run_id = self._previous_snapshot(state.run_id)
        change_report = self._daily_change_report(
            snapshot, previous_snapshot, rules, previous_run_id
        )
        performance_report = self._collector_performance_report(
            snapshot, previous_snapshot, performance_rules
        )
        alerts = self._collector_alerts(performance_report, snapshot, performance_rules)
        monitoring_dashboard = self._manager_daily_monitoring_dashboard(
            snapshot,
            previous_snapshot,
            change_report,
            performance_report,
            alerts,
            imported_count=len(artifacts.get("raw_rows", [])),
            valid_count=state.processed_debt_records,
        )
        summary = {
            "date": date.today().isoformat(),
            "run_id": state.run_id,
            "snapshot_count": len(snapshot),
            "previous_run_id": previous_run_id,
            "total_debt": round(sum(record["debt_amount"] for record in snapshot), 2),
            "queue_counts": self._count_by(snapshot, "last_action_queue"),
            "recommendation_only": True,
        }
        self._write_json(state.run_id, "daily_snapshot.json", snapshot)
        self._write_json(state.run_id, "snapshot_summary.json", summary)
        self._write_json(state.run_id, "daily_change_report.json", change_report)
        self._write_json(state.run_id, "collector_performance_report.json", performance_report)
        self._write_json(state.run_id, "collector_alerts.json", alerts)
        self._write_json(
            state.run_id, "manager_daily_monitoring_dashboard.json", monitoring_dashboard
        )
        return {
            "daily_snapshot": snapshot,
            "snapshot_summary": summary,
            "daily_change_report": change_report,
            "collector_performance_report": performance_report,
            "collector_alerts": alerts,
            "manager_daily_monitoring_dashboard": monitoring_dashboard,
        }

    def _generate_manager_briefing(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        recs = artifacts.get("recommendations", [])
        total = round(sum(r["debt_amount"] for r in recs), 2)
        forecast = round(
            sum(r["debt_amount"] * r["estimated_success_probability"] for r in recs), 2
        )
        briefing = {
            "date": date.today().isoformat(),
            "today_total_target": total,
            "total_collectible_amount": forecast,
            "top_priority_cases": [r for r in recs if r["priority"] == "high"][:10],
            "overdue_follow_ups": [r for r in recs if r["recommended_action"] == "FIELD_VISIT"],
            "region_comparison": self._sum_by(recs, "region"),
            "collector_comparison": self._sum_by(recs, "collector"),
            "expected_collection_forecast": forecast,
            "recommended_management_actions": self._management_actions(recs),
        }
        self._write_json(state.run_id, "manager_morning_briefing.json", briefing)
        self._mark_seen(artifacts.get("files", []))
        return {"manager_briefing": briefing}

    def _generate_executive_intelligence(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        recs = artifacts.get("recommendations", [])
        scored = artifacts.get("scored_records", [])
        region_stats = self._performance_stats(recs, "region")
        collector_stats = self._performance_stats(recs, "collector")
        district_stats = self._district_stats(scored)
        total_debt = round(sum(r["debt_amount"] for r in recs), 2)
        expected_revenue = round(
            sum(r["debt_amount"] * r["estimated_success_probability"] for r in recs), 2
        )
        end_of_month_revenue = self._end_of_month_forecast(expected_revenue)
        efficiency = self._collection_efficiency_score(total_debt, expected_revenue, recs)
        health = self._municipality_health_score(efficiency, region_stats, collector_stats, scored)

        recommendations = []
        recommendations.extend(self._weak_region_recommendations(region_stats))
        recommendations.extend(self._weak_collector_recommendations(collector_stats))
        recommendations.extend(self._uncollectible_debt_recommendations(recs))
        recommendations.extend(self._collector_intervention_recommendations(collector_stats))
        recommendations.extend(self._special_negotiation_recommendations(recs))
        recommendations.extend(self._business_closure_recommendations(scored))
        recommendations.extend(self._commission_77_recommendations(scored))
        recommendations.extend(self._field_campaign_recommendations(district_stats))
        recommendations.extend(self._seasonal_pattern_recommendations(recs))
        recommendations.extend(
            self._abnormal_behavior_recommendations(region_stats, collector_stats, scored)
        )
        recommendations.extend(
            self._daily_strategic_recommendations(health, efficiency, end_of_month_revenue, recs)
        )

        dashboard = {
            "date": date.today().isoformat(),
            "executive_kpis": {
                "open_debt_amount": total_debt,
                "expected_collectible_amount": expected_revenue,
                "predicted_end_of_month_revenue": end_of_month_revenue,
                "collection_efficiency_score": efficiency,
                "municipality_health_score": health,
                "high_priority_case_count": sum(1 for r in recs if r["priority"] == "high"),
                "probably_uncollectible_amount": round(
                    sum(
                        r["debt_amount"] for r in recs if r["estimated_success_probability"] <= 0.25
                    ),
                    2,
                ),
                "executive_recommendation_count": len(recommendations),
            },
            "weak_performing_regions": [
                name for name, stat in region_stats.items() if stat["efficiency"] < 45
            ],
            "weak_performing_collectors": [
                name for name, stat in collector_stats.items() if stat["efficiency"] < 45
            ],
            "districts_requiring_field_campaigns": [
                name
                for name, stat in district_stats.items()
                if stat["high_priority_count"] >= 3 or stat["debt_amount"] >= 10000
            ],
            "daily_strategic_recommendations": [
                r for r in recommendations if r["category"] == "daily_strategy"
            ],
        }
        intelligence = {
            "mayor_dashboard": dashboard,
            "region_performance": region_stats,
            "collector_performance": collector_stats,
            "district_performance": district_stats,
            "executive_recommendations": recommendations,
        }
        state.executive_recommendations_generated = len(recommendations)
        self._write_json(state.run_id, "mayor_dashboard.json", dashboard)
        self._write_json(state.run_id, "executive_decision_engine.json", intelligence)
        self._store_executive_recommendations(state.run_id, recommendations)
        return {"executive_intelligence": intelligence, "mayor_dashboard": dashboard}

    def _generate_field_activity_attribution(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        previous_snapshot, _ = self._previous_snapshot(state.run_id)
        activities = self._load_field_activity_log()
        attribution = FieldActivityAttributionEngine(self.workspace).build(
            run_id=state.run_id,
            current_records=artifacts.get("daily_snapshot", []),
            previous_records=previous_snapshot,
            activities=activities,
        )
        for filename in [
            "field_activity_attribution_report.json",
            "collector_credit_report.json",
            "unattributed_collections.json",
            "field_activity_quality_report.json",
            "manager_field_performance_dashboard.json",
        ]:
            key = filename.removesuffix(".json")
            self._write_json(state.run_id, filename, attribution[key])
        return {"field_activity_attribution": attribution}

    def _load_field_activity_log(self) -> list[dict[str, Any]]:
        path = self.inbox_dir / "field_activity_log.json"
        if not path.exists():
            return []
        payload = json.loads(path.read_text())
        if isinstance(payload, dict):
            return list(payload.get("activities", []))
        return list(payload)

    def _generate_financial_intelligence(
        self, state: RunState, artifacts: dict[str, Any]
    ) -> dict[str, Any]:
        previous_snapshot, _ = self._previous_snapshot(state.run_id)
        dashboard = FinancialIntelligenceDashboard(self.workspace).build(
            run_id=state.run_id,
            current_snapshot=artifacts.get("daily_snapshot", []),
            previous_snapshot=previous_snapshot,
        )
        self._write_json(state.run_id, "financial_intelligence_dashboard.json", dashboard)
        return {"financial_intelligence_dashboard": dashboard}

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open(newline="") as handle:
            return list(csv.DictReader(handle))

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        sheet = load_workbook(path, read_only=True, data_only=True).active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell or "") for cell in rows[0]]
        return [
            dict(zip(headers, row, strict=False))
            for row in rows[1:]
            if any(cell is not None for cell in row)
        ]

    def _save_state(self, state: RunState) -> None:
        self._write_json(state.run_id, "run_state.json", asdict(state))

    def _save_step_artifacts(self, run_id: str, step: str, artifacts: dict[str, Any]) -> None:
        self._write_json(run_id, f"{step}.json", artifacts)

    def _load_step_artifacts(self, run_id: str, step: str) -> dict[str, Any]:
        path = self.runs_dir / run_id / f"{step}.json"
        return json.loads(path.read_text()) if path.exists() else {}

    def _write_json(self, run_id: str, name: str, data: Any) -> None:
        run_dir = self.runs_dir / run_id
        run_dir.mkdir(parents=True, exist_ok=True)
        (run_dir / name).write_text(json.dumps(data, indent=2, sort_keys=True))

    def _audit(self, run_id: str, event: str, details: dict[str, Any]) -> None:
        entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "run_id": run_id,
            "event": event,
            "details": details,
        }
        with (self.audit_dir / f"{date.today().isoformat()}.jsonl").open("a") as handle:
            handle.write(json.dumps(entry, sort_keys=True) + "\n")

    def _seen_file_hashes(self) -> set[str]:
        path = self.state_dir / "seen_files.json"
        return set(json.loads(path.read_text())) if path.exists() else set()

    def _mark_seen(self, files: list[dict[str, Any]]) -> None:
        seen = self._seen_file_hashes()
        seen.update(file["sha256"] for file in files)
        (self.state_dir / "seen_files.json").write_text(json.dumps(sorted(seen), indent=2))

    def _store_executive_recommendations(
        self, run_id: str, recommendations: list[dict[str, Any]]
    ) -> None:
        learning_path = self.state_dir / "executive_recommendations.jsonl"
        with learning_path.open("a") as handle:
            for recommendation in recommendations:
                entry = {
                    "run_id": run_id,
                    "stored_at": datetime.now(timezone.utc).isoformat(),
                    **recommendation,
                }
                handle.write(json.dumps(entry, sort_keys=True) + "\n")

    @staticmethod
    def _performance_stats(
        records: list[dict[str, Any]], field_name: str
    ) -> dict[str, dict[str, Any]]:
        stats: dict[str, dict[str, Any]] = {}
        for record in records:
            bucket = stats.setdefault(
                record[field_name],
                {"case_count": 0, "debt_amount": 0.0, "forecast": 0.0, "high_priority_count": 0},
            )
            bucket["case_count"] += 1
            bucket["debt_amount"] = round(bucket["debt_amount"] + record["debt_amount"], 2)
            bucket["forecast"] = round(
                bucket["forecast"]
                + record["debt_amount"] * record["estimated_success_probability"],
                2,
            )
            bucket["high_priority_count"] += 1 if record["priority"] == "high" else 0
        for bucket in stats.values():
            bucket["efficiency"] = round(
                (bucket["forecast"] / bucket["debt_amount"] * 100)
                if bucket["debt_amount"]
                else 100,
                2,
            )
        return stats

    @staticmethod
    def _district_stats(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        stats: dict[str, dict[str, Any]] = {}
        for record in records:
            district = CollectionAgent._district_from_record(record)
            bucket = stats.setdefault(
                district, {"case_count": 0, "debt_amount": 0.0, "high_priority_count": 0}
            )
            bucket["case_count"] += 1
            bucket["debt_amount"] = round(bucket["debt_amount"] + record["debt_amount"], 2)
            bucket["high_priority_count"] += 1 if record.get("priority") == "high" else 0
        return stats

    @staticmethod
    def _district_from_record(record: dict[str, Any]) -> str:
        return str(record.get("district") or record.get("region") or "Unassigned")

    @staticmethod
    def _end_of_month_forecast(expected_revenue: float) -> float:
        today = date.today()
        elapsed_ratio = max(today.day / 30, 0.1)
        return round(expected_revenue / elapsed_ratio, 2)

    @staticmethod
    def _collection_efficiency_score(
        total_debt: float, expected_revenue: float, records: list[dict[str, Any]]
    ) -> float:
        if not records or total_debt <= 0:
            return 100.0
        high_priority_penalty = (
            sum(1 for r in records if r["priority"] == "high") / len(records) * 15
        )
        return round(
            max(0, min(100, (expected_revenue / total_debt * 100) - high_priority_penalty)), 2
        )

    @staticmethod
    def _municipality_health_score(
        efficiency: float,
        region_stats: dict[str, dict[str, Any]],
        collector_stats: dict[str, dict[str, Any]],
        records: list[dict[str, Any]],
    ) -> float:
        weak_regions = sum(1 for stat in region_stats.values() if stat["efficiency"] < 45)
        weak_collectors = sum(1 for stat in collector_stats.values() if stat["efficiency"] < 45)
        abnormal_penalty = sum(1 for r in records if r.get("debt_amount", 0) >= 25000) * 3
        return round(
            max(
                0, min(100, efficiency - weak_regions * 4 - weak_collectors * 3 - abnormal_penalty)
            ),
            2,
        )

    @staticmethod
    def _executive_recommendation(
        category: str,
        title: str,
        severity: str,
        opportunity: str,
        action: str,
        reasoning: str,
        evidence: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "id": hashlib.sha256(
                f"{category}|{title}|{json.dumps(evidence, sort_keys=True)}".encode()
            ).hexdigest()[:16],
            "category": category,
            "title": title,
            "severity": severity,
            "management_opportunity": opportunity,
            "recommended_action": action,
            "ai_reasoning": reasoning,
            "evidence": evidence,
        }

    def _weak_region_recommendations(
        self, stats: dict[str, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "weak_region",
                f"Weak-performing region: {name}",
                "high",
                "Recover revenue by reallocating field capacity to underperforming geography.",
                "Review district barriers, add supervisor ride-alongs, and rebalance collector assignments today.",
                f"AI reasoning: {name} has an efficiency score of {stat['efficiency']} with {stat['high_priority_count']} high-priority cases, indicating the region is converting less debt into expected revenue than the municipal baseline.",
                {"region": name, **stat},
            )
            for name, stat in stats.items()
            if stat["efficiency"] < 45
        ]

    def _weak_collector_recommendations(
        self, stats: dict[str, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "weak_collector",
                f"Weak-performing collector: {name}",
                "medium",
                "Improve team productivity through targeted coaching.",
                "Audit the collector queue, review contact scripts, and pair with a top performer for the next route.",
                f"AI reasoning: {name} owns a portfolio with {stat['efficiency']} efficiency and {stat['high_priority_count']} high-priority cases, which suggests intervention may increase near-term collections.",
                {"collector": name, **stat},
            )
            for name, stat in stats.items()
            if stat["efficiency"] < 45
        ]

    def _uncollectible_debt_recommendations(
        self, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "probably_uncollectible",
                f"Probably uncollectible debt: {r['debtor']}",
                "high",
                "Reduce wasted effort by separating low-probability debt from daily collector routes.",
                "Move to senior review for write-down, legal escalation, or settlement authority.",
                f"AI reasoning: Success probability is {r['estimated_success_probability']} while debt is ${r['debt_amount']:.2f}; this combination indicates normal outreach is unlikely to recover the balance.",
                {
                    "record_id": r["record_id"],
                    "debtor": r["debtor"],
                    "probability": r["estimated_success_probability"],
                    "debt_amount": r["debt_amount"],
                },
            )
            for r in records
            if r["estimated_success_probability"] <= 0.25
        ]

    def _collector_intervention_recommendations(
        self, stats: dict[str, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "collector_intervention",
                f"Collector intervention required: {name}",
                "high",
                "Prevent a portfolio from becoming structurally delinquent.",
                "Manager should inspect case notes, approve a route plan, and set a same-week recovery target.",
                f"AI reasoning: {name} combines low efficiency ({stat['efficiency']}) with at least three high-priority cases, a pattern that requires direct management intervention.",
                {"collector": name, **stat},
            )
            for name, stat in stats.items()
            if stat["efficiency"] < 55 and stat["high_priority_count"] >= 3
        ]

    def _special_negotiation_recommendations(
        self, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "special_negotiation",
                f"Special negotiation candidate: {r['debtor']}",
                "medium",
                "Increase recoveries through structured payment plans before legal costs rise.",
                "Offer a manager-approved payment plan or settlement window within 48 hours.",
                f"AI reasoning: The debt is high (${r['debt_amount']:.2f}) but probability remains {r['estimated_success_probability']}, making negotiation more attractive than routine reminders.",
                {
                    "record_id": r["record_id"],
                    "debtor": r["debtor"],
                    "debt_amount": r["debt_amount"],
                    "probability": r["estimated_success_probability"],
                },
            )
            for r in records
            if r["debt_amount"] >= 5000 and 0.25 < r["estimated_success_probability"] <= 0.6
        ]

    def _business_closure_recommendations(
        self, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        candidates = [
            r
            for r in records
            if "business" in str(r.get("debtor", "")).lower() or r.get("debt_amount", 0) >= 15000
        ]
        return [
            self._executive_recommendation(
                "business_closure_risk",
                f"Business closure risk: {r['debtor']}",
                "high",
                "Protect municipal revenue before the taxpayer exits operations.",
                "Schedule immediate business outreach and verify operating status through field inspection.",
                f"AI reasoning: {r['debtor']} has a large balance (${r['debt_amount']:.2f}) and high collection pressure, which can indicate closure or insolvency risk if not addressed quickly.",
                {
                    "record_id": r["record_id"],
                    "debtor": r["debtor"],
                    "debt_amount": r["debt_amount"],
                    "score": r.get("score"),
                },
            )
            for r in candidates
            if r.get("score", 0) >= 80
        ]

    def _commission_77_recommendations(self, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "commission_77",
                f"Move legal case to Commission 77: {r['debtor']}",
                "critical",
                "Accelerate legally actionable debts with high municipal exposure.",
                "Prepare Commission 77 packet immediately and remove from routine queue after manager approval.",
                f"AI reasoning: The case combines high debt (${r['debt_amount']:.2f}) and maximum urgency score ({r.get('score')}), meeting the immediate escalation threshold.",
                {
                    "record_id": r["record_id"],
                    "debtor": r["debtor"],
                    "debt_amount": r["debt_amount"],
                    "score": r.get("score"),
                },
            )
            for r in records
            if r.get("score", 0) >= 95 and r["debt_amount"] >= 10000
        ]

    def _field_campaign_recommendations(
        self, stats: dict[str, dict[str, Any]]
    ) -> list[dict[str, Any]]:
        return [
            self._executive_recommendation(
                "field_campaign",
                f"Field campaign needed in {name}",
                "high",
                "Capture concentrated revenue through coordinated field operations.",
                "Launch a district field campaign with route batching and supervisor review by end of day.",
                f"AI reasoning: {name} has {stat['high_priority_count']} high-priority cases and ${stat['debt_amount']:.2f} in open debt, indicating geographic concentration.",
                {"district": name, **stat},
            )
            for name, stat in stats.items()
            if stat["high_priority_count"] >= 3 or stat["debt_amount"] >= 10000
        ]

    def _seasonal_pattern_recommendations(
        self, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        current_month = date.today().month
        if current_month in {1, 2, 11, 12} and records:
            amount = round(sum(r["debt_amount"] for r in records), 2)
            return [
                self._executive_recommendation(
                    "seasonal_pattern",
                    "Seasonal collection pressure detected",
                    "medium",
                    "Adjust staffing before seasonal payment behavior reduces cash recovery.",
                    "Increase early-month reminders and prioritize negotiated settlements during the seasonal risk window.",
                    f"AI reasoning: Month {current_month} historically behaves like a seasonal risk window; the current queue contains ${amount:.2f}, so earlier outreach can protect cash flow.",
                    {"month": current_month, "debt_amount": amount, "case_count": len(records)},
                )
            ]
        return []

    def _abnormal_behavior_recommendations(
        self,
        region_stats: dict[str, dict[str, Any]],
        collector_stats: dict[str, dict[str, Any]],
        records: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        recommendations = []
        for record in records:
            if record.get("debt_amount", 0) >= 25000:
                recommendations.append(
                    self._executive_recommendation(
                        "abnormal_behavior",
                        f"Abnormal high-value debt: {record['debtor']}",
                        "critical",
                        "Prevent one outlier from distorting collection performance and risk exposure.",
                        "Assign executive review and verify taxpayer status, balance accuracy, and legal readiness today.",
                        f"AI reasoning: ${record['debt_amount']:.2f} is materially larger than normal daily cases, making it an abnormal exposure requiring validation.",
                        {
                            "record_id": record["record_id"],
                            "debtor": record["debtor"],
                            "debt_amount": record["debt_amount"],
                        },
                    )
                )
        return recommendations

    def _daily_strategic_recommendations(
        self, health: float, efficiency: float, forecast: float, records: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        severity = "high" if health < 50 or efficiency < 45 else "medium"
        return [
            self._executive_recommendation(
                "daily_strategy",
                "Daily executive collection strategy",
                severity,
                "Align management attention with the highest revenue and operational risks.",
                "Hold a 15-minute morning command meeting, confirm high-priority owners, and track forecast variance by 3:00 PM.",
                f"AI reasoning: Municipality health is {health}, collection efficiency is {efficiency}, and end-of-month revenue is forecast at ${forecast:.2f}; executive coordination is needed to protect the forecast across {len(records)} active cases.",
                {
                    "municipality_health_score": health,
                    "collection_efficiency_score": efficiency,
                    "predicted_end_of_month_revenue": forecast,
                    "case_count": len(records),
                },
            )
        ]

    @staticmethod
    def _normalize_key(key: Any) -> str:
        normalized = (
            str(key)
            .strip()
            .lower()
            .replace("ي", "ی")
            .replace("ك", "ک")
            .replace(" ", "_")
            .replace("-", "_")
            .replace("\n", "_")
        )
        while "__" in normalized:
            normalized = normalized.replace("__", "_")
        return normalized.strip("_")

    @staticmethod
    def _apply_schema_mapping(row: dict[str, Any]) -> None:
        for target, aliases in REGION_1_SCHEMA_ALIASES.items():
            if row.get(target):
                continue
            for alias in aliases:
                value = row.get(alias)
                if value not in (None, ""):
                    row[target] = value
                    break

    @staticmethod
    def _municipal_action(record: dict[str, Any]) -> str:
        status = str(record.get("status", "")).lower()
        if status in {"closed", "paid", "resolved", "ignore"}:
            return "IGNORE"
        if status in {"follow_up", "follow-up", "follow up"}:
            return "FOLLOW_UP"
        if status in {"legal", "legal_review", "legal review"}:
            return "LEGAL"
        if record.get("priority") == "high":
            return "FIELD_VISIT"
        if record.get("priority") == "medium":
            return "CALL"
        return "NOTICE"

    def _manager_assignment_dashboard(
        self, queues: dict[str, list[dict[str, Any]]]
    ) -> dict[str, Any]:
        records = [record for queue in queues.values() for record in queue]
        return {
            "date": date.today().isoformat(),
            "recommendation_only": True,
            "queue_counts": {action: len(queue) for action, queue in queues.items()},
            "generated_outputs": MUNICIPAL_ACTION_QUEUES,
            "summaries": {
                "by_region": self._queue_summary(records, "region"),
                "by_zone": self._queue_summary(records, "zone"),
                "by_collector": self._queue_summary(records, "collector"),
                "by_priority_score": self._queue_summary(records, "priority_score"),
            },
            "unassigned_records": [
                record for record in records if not record.get("recommended_action")
            ],
            "review_notes": [
                "Recommendation-only queues generated; no messages or actions were sent."
            ],
        }

    @staticmethod
    def _queue_summary(records: list[dict[str, Any]], field_name: str) -> dict[str, dict[str, Any]]:
        summary: dict[str, dict[str, Any]] = {}
        for record in records:
            key = str(record.get(field_name) or "Unassigned")
            bucket = summary.setdefault(
                key,
                {
                    "case_count": 0,
                    "debt_amount": 0.0,
                    "queues": {action: 0 for action in MUNICIPAL_ACTION_QUEUES},
                },
            )
            bucket["case_count"] += 1
            bucket["debt_amount"] = round(bucket["debt_amount"] + record.get("debt_amount", 0), 2)
            bucket["queues"][record["recommended_action"]] += 1
        return summary

    def _snapshot_record(
        self, record: dict[str, Any], recommendation: dict[str, Any] | None, run_id: str
    ) -> dict[str, Any]:
        return {
            "debtor_id": record["debtor_id"],
            "debtor_name": record["debtor"],
            "address": record["address"],
            "mobile": record.get("mobile"),
            "debt_amount": record["debt_amount"],
            "region": record.get("region") or "Unassigned",
            "zone": record.get("zone") or record.get("region") or "Unassigned",
            "business_category": record.get("business_category"),
            "legal_status": record.get("legal_status") or record.get("status") or "open",
            "last_action_queue": (recommendation or {}).get("recommended_action", "IGNORE"),
            "assigned_collector": record.get("collector") or "Unassigned",
            "priority_score": record.get("score", 0),
            "source_file_name": record.get("source_file", ""),
            "import_date": date.today().isoformat(),
            "run_id": run_id,
        }

    def _previous_snapshot(self, current_run_id: str) -> tuple[list[dict[str, Any]], str | None]:
        candidates: list[tuple[str, str, Path]] = []
        for run_dir in self.runs_dir.iterdir():
            if not run_dir.is_dir() or run_dir.name == current_run_id:
                continue
            snapshot_path = run_dir / "daily_snapshot.json"
            state_path = run_dir / "run_state.json"
            if not snapshot_path.exists():
                continue
            start_time = ""
            if state_path.exists():
                start_time = json.loads(state_path.read_text()).get("start_time", "")
            candidates.append((start_time, run_dir.name, snapshot_path))
        if not candidates:
            return [], None
        _, run_id, path = sorted(candidates)[-1]
        return json.loads(path.read_text()), run_id

    def _daily_change_report(
        self,
        snapshot: list[dict[str, Any]],
        previous_snapshot: list[dict[str, Any]],
        rules: dict[str, Any],
        previous_run_id: str | None,
    ) -> dict[str, Any]:
        current = {record["debtor_id"]: record for record in snapshot}
        previous = {record["debtor_id"]: record for record in previous_snapshot}
        shared_ids = set(current) & set(previous)
        min_change = float(rules.get("minimum_debt_change", 0))
        paid_threshold = float(rules.get("paid_debt_threshold", 0))
        cleared_statuses = {
            str(status).lower() for status in rules.get("cleared_legal_statuses", [])
        }

        increased, decreased, unchanged = [], [], []
        changed_address_or_mobile, changed_legal_status, changed_action_queue = [], [], []
        fully_paid_or_cleared = []
        for debtor_id in sorted(shared_ids):
            before = previous[debtor_id]
            after = current[debtor_id]
            delta = round(after["debt_amount"] - before["debt_amount"], 2)
            if delta > min_change:
                increased.append(
                    {"debtor_id": debtor_id, "before": before, "after": after, "delta": delta}
                )
            elif delta < -min_change:
                decreased.append(
                    {"debtor_id": debtor_id, "before": before, "after": after, "delta": delta}
                )
            else:
                unchanged.append({"debtor_id": debtor_id, "record": after})
            if before.get("address") != after.get("address") or before.get("mobile") != after.get(
                "mobile"
            ):
                changed_address_or_mobile.append(
                    {"debtor_id": debtor_id, "before": before, "after": after}
                )
            if before.get("legal_status") != after.get("legal_status"):
                changed_legal_status.append(
                    {"debtor_id": debtor_id, "before": before, "after": after}
                )
            if before.get("last_action_queue") != after.get("last_action_queue"):
                changed_action_queue.append(
                    {"debtor_id": debtor_id, "before": before, "after": after}
                )
            if (
                after["debt_amount"] <= paid_threshold
                or str(after.get("legal_status", "")).lower() in cleared_statuses
            ):
                fully_paid_or_cleared.append(after)

        removed = [previous[debtor_id] for debtor_id in sorted(set(previous) - set(current))]
        return {
            "date": date.today().isoformat(),
            "previous_run_id": previous_run_id,
            "recommendation_only": True,
            "new_debtors": [
                current[debtor_id] for debtor_id in sorted(set(current) - set(previous))
            ],
            "removed_debtors": removed,
            "debt_amount_increased": increased,
            "debt_amount_decreased": decreased,
            "fully_paid_or_cleared_debtors": fully_paid_or_cleared,
            "unchanged_debtors": unchanged,
            "changed_address_or_mobile": changed_address_or_mobile,
            "changed_legal_status": changed_legal_status,
            "changed_action_queue": changed_action_queue,
            "queue_count_changes": self._queue_count_changes(snapshot, previous_snapshot),
        }

    def _collector_performance_report(
        self,
        snapshot: list[dict[str, Any]],
        previous_snapshot: list[dict[str, Any]],
        rules: dict[str, Any],
    ) -> dict[str, Any]:
        current = {record["debtor_id"]: record for record in snapshot}
        rows: dict[str, dict[str, Any]] = {}
        for before in previous_snapshot:
            collector = before.get("assigned_collector") or "Unassigned"
            row = rows.setdefault(collector, self._empty_collector_performance(collector))
            row["assigned_case_count"] += 1
            after = current.get(before["debtor_id"])
            if after is None:
                row["paid_or_cleared_cases"] += 1
                row["cases_improved"] += 1
                row["debt_reduced_amount"] = round(
                    row["debt_reduced_amount"] + before["debt_amount"], 2
                )
                continue
            delta = round(after["debt_amount"] - before["debt_amount"], 2)
            action = before.get("last_action_queue", "IGNORE")
            if delta < 0:
                row["cases_improved"] += 1
                row["debt_reduced_amount"] = round(row["debt_reduced_amount"] + abs(delta), 2)
                self._increment_impact(row, action)
            elif delta > 0:
                row["cases_worsened"] += 1
                row["debt_increased_amount"] = round(row["debt_increased_amount"] + delta, 2)
            else:
                row["cases_unchanged"] += 1
                if action in {"FOLLOW_UP", "FIELD_VISIT", "NOTICE", "CALL"}:
                    row["follow_up_overdue_cases"] += 1
        for row in rows.values():
            row["performance_score"] = self._performance_score(row, rules)
        ranking = sorted(rows.values(), key=lambda row: row["performance_score"], reverse=True)
        return {
            "date": date.today().isoformat(),
            "recommendation_only": True,
            "collectors": ranking,
        }

    def _collector_alerts(
        self,
        performance_report: dict[str, Any],
        snapshot: list[dict[str, Any]],
        rules: dict[str, Any],
    ) -> dict[str, Any]:
        alerts = []
        high_priority_score = int(rules.get("high_priority_ignored_score", 80))
        for row in performance_report.get("collectors", []):
            assigned = max(row["assigned_case_count"], 1)
            unchanged_ratio = row["cases_unchanged"] / assigned
            if row["cases_unchanged"] >= rules.get("unchanged_many_threshold", 10):
                alerts.append(
                    {
                        "collector": row["collector"],
                        "alert": "collector_has_many_unchanged_cases",
                        "evidence": row,
                    }
                )
            if unchanged_ratio >= rules.get("unchanged_ratio_alert", 0.7):
                alerts.append(
                    {
                        "collector": row["collector"],
                        "alert": "reported_work_without_debt_change",
                        "evidence": row,
                    }
                )
            if row["field_visit_impact"] == 0 and row["follow_up_overdue_cases"] >= rules.get(
                "repeated_visit_no_progress_threshold", 2
            ):
                alerts.append(
                    {
                        "collector": row["collector"],
                        "alert": "repeated_visits_with_no_progress",
                        "evidence": row,
                    }
                )
            if row["notice_impact"] == 0 and row["follow_up_overdue_cases"] >= rules.get(
                "notice_no_payment_threshold", 5
            ):
                alerts.append(
                    {
                        "collector": row["collector"],
                        "alert": "notices_issued_no_payment_movement",
                        "evidence": row,
                    }
                )
        ignored = [
            record
            for record in snapshot
            if record.get("priority_score", 0) >= high_priority_score
            and record.get("last_action_queue") == "IGNORE"
        ]
        if ignored:
            alerts.append(
                {
                    "collector": "Unassigned",
                    "alert": "high_priority_cases_ignored",
                    "evidence": ignored,
                }
            )
        return {"date": date.today().isoformat(), "recommendation_only": True, "alerts": alerts}

    def _manager_daily_monitoring_dashboard(
        self,
        snapshot: list[dict[str, Any]],
        previous_snapshot: list[dict[str, Any]],
        change_report: dict[str, Any],
        performance_report: dict[str, Any],
        alerts: dict[str, Any],
        imported_count: int,
        valid_count: int,
    ) -> dict[str, Any]:
        yesterday_debt = round(sum(record["debt_amount"] for record in previous_snapshot), 2)
        today_debt = round(sum(record["debt_amount"] for record in snapshot), 2)
        worsened = change_report["debt_amount_increased"]
        improved = change_report["debt_amount_decreased"]
        return {
            "date": date.today().isoformat(),
            "recommendation_only": True,
            "total_imported_records_today": imported_count,
            "total_valid_records": valid_count,
            "total_new_debtors": len(change_report["new_debtors"]),
            "total_removed_or_paid_debtors": len(change_report["removed_debtors"])
            + len(change_report["fully_paid_or_cleared_debtors"]),
            "total_debt_yesterday": yesterday_debt,
            "total_debt_today": today_debt,
            "net_debt_change": round(today_debt - yesterday_debt, 2),
            "queue_count_changes": change_report["queue_count_changes"],
            "collector_ranking": performance_report.get("collectors", []),
            "worst_collector_risks": alerts.get("alerts", [])[:10],
            "top_improved_cases": sorted(improved, key=lambda item: item["delta"])[:10],
            "top_worsened_cases": sorted(worsened, key=lambda item: item["delta"], reverse=True)[
                :10
            ],
            "recommended_manager_actions": self._monitoring_manager_actions(change_report, alerts),
        }

    @staticmethod
    def _empty_collector_performance(collector: str) -> dict[str, Any]:
        return {
            "collector": collector,
            "assigned_case_count": 0,
            "cases_improved": 0,
            "cases_unchanged": 0,
            "cases_worsened": 0,
            "debt_reduced_amount": 0.0,
            "debt_increased_amount": 0.0,
            "paid_or_cleared_cases": 0,
            "follow_up_overdue_cases": 0,
            "field_visit_impact": 0,
            "notice_impact": 0,
            "call_impact": 0,
            "performance_score": 0,
        }

    @staticmethod
    def _increment_impact(row: dict[str, Any], action: str) -> None:
        if action == "FIELD_VISIT":
            row["field_visit_impact"] += 1
        elif action == "NOTICE":
            row["notice_impact"] += 1
        elif action == "CALL":
            row["call_impact"] += 1

    @staticmethod
    def _performance_score(row: dict[str, Any], rules: dict[str, Any]) -> int:
        weights = rules.get("score_weights", {})
        score = float(weights.get("base", 70))
        score += row["cases_improved"] * float(weights.get("improved", 8))
        score += row["paid_or_cleared_cases"] * float(weights.get("paid", 10))
        score += row["cases_worsened"] * float(weights.get("worsened", -12))
        score += row["cases_unchanged"] * float(weights.get("unchanged", -2))
        score += row["follow_up_overdue_cases"] * float(weights.get("overdue", -5))
        return int(max(0, min(100, round(score))))

    @staticmethod
    def _queue_count_changes(
        snapshot: list[dict[str, Any]], previous_snapshot: list[dict[str, Any]]
    ) -> dict[str, dict[str, int]]:
        today = CollectionAgent._count_by(snapshot, "last_action_queue")
        yesterday = CollectionAgent._count_by(previous_snapshot, "last_action_queue")
        return {
            action: {
                "yesterday": yesterday.get(action, 0),
                "today": today.get(action, 0),
                "delta": today.get(action, 0) - yesterday.get(action, 0),
            }
            for action in MUNICIPAL_ACTION_QUEUES
        }

    @staticmethod
    def _monitoring_manager_actions(
        change_report: dict[str, Any], alerts: dict[str, Any]
    ) -> list[str]:
        actions = ["Review daily snapshot deltas before approving enforcement work."]
        if change_report["debt_amount_increased"]:
            actions.append(
                "Audit the largest debt increases and confirm balances against source files."
            )
        if alerts.get("alerts"):
            actions.append("Review collector alerts and require evidence for no-progress cases.")
        if change_report["new_debtors"]:
            actions.append(
                "Assign new debtors to the appropriate recommendation-only action queues."
            )
        return actions

    def _load_json_config(self, filename: str, defaults: dict[str, Any]) -> dict[str, Any]:
        path = Path("config") / filename
        rules = json.loads(json.dumps(defaults))
        if path.exists():
            rules.update(json.loads(path.read_text()))
        return rules

    @staticmethod
    def _clean_value(value: Any) -> Any:
        return value.strip() if isinstance(value, str) else value

    @staticmethod
    def _money(value: Any) -> float:
        if value in (None, ""):
            return 0.0
        return round(float(str(value).replace("$", "").replace(",", "")), 2)

    @staticmethod
    def _record_id(row: dict[str, Any]) -> str:
        basis = f"{row.get('debtor')}|{row.get('address')}|{row.get('debt_amount')}|{row.get('source_file')}|{row.get('row_number')}"
        return hashlib.sha256(basis.encode()).hexdigest()[:16]

    @staticmethod
    def _debtor_id(row: dict[str, Any]) -> str:
        explicit_id = row.get("debtor_id")
        if explicit_id not in (None, ""):
            if str(row.get("source_file", "")).lower().endswith(".xlsx"):
                basis = (
                    f"{explicit_id}|{row.get('debtor')}|{row.get('source_file')}|"
                    f"{row.get('row_number')}"
                )
                return hashlib.sha256(basis.encode()).hexdigest()[:16]
            return str(explicit_id)
        basis = f"{row.get('debtor')}|{row.get('mobile')}|{row.get('address')}"
        return hashlib.sha256(basis.encode()).hexdigest()[:16]

    @staticmethod
    def _count_by(records: list[dict[str, Any]], field_name: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for record in records:
            key = str(record.get(field_name) or "Unassigned")
            counts[key] = counts.get(key, 0) + 1
        return counts

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _days_since(value: str | None, today: date) -> int:
        if not value:
            return 0
        try:
            return (today - datetime.fromisoformat(str(value)).date()).days
        except ValueError:
            return 0

    @staticmethod
    def _sum_by(
        records: list[dict[str, Any]], field_name: str
    ) -> dict[str, dict[str, float | int]]:
        output: dict[str, dict[str, float | int]] = {}
        for record in records:
            bucket = output.setdefault(
                record[field_name], {"case_count": 0, "debt_amount": 0.0, "forecast": 0.0}
            )
            bucket["case_count"] += 1
            bucket["debt_amount"] = round(float(bucket["debt_amount"]) + record["debt_amount"], 2)
            bucket["forecast"] = round(
                float(bucket["forecast"])
                + record["debt_amount"] * record["estimated_success_probability"],
                2,
            )
        return output

    @staticmethod
    def _management_actions(records: list[dict[str, Any]]) -> list[str]:
        high = sum(1 for r in records if r["priority"] == "high")
        actions = ["Review collector queues by 9:00 AM and confirm field capacity."]
        if high:
            actions.append(f"Assign same-day outreach for {high} high-priority cases.")
        if not records:
            actions.append("No new files were detected; verify upload pipeline health.")
        return actions


def schedule_daily(workspace: str, hour: int = 6, minute: int = 0) -> None:
    """Run forever and execute once per UTC day at the configured time."""
    last_run: str | None = None
    agent = CollectionAgent(workspace)
    while True:  # pragma: no cover - long-running scheduler loop
        now = datetime.now(timezone.utc)
        if now.hour == hour and now.minute == minute and last_run != now.date().isoformat():
            agent.run(mode="scheduled")
            last_run = now.date().isoformat()
        time.sleep(30)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Autonomous Collection Agent v1")
    parser.add_argument("command", choices=["run", "schedule"])
    parser.add_argument("--workspace", default="var/collection_agent")
    parser.add_argument("--resume-run-id")
    parser.add_argument("--hour", type=int, default=6)
    parser.add_argument("--minute", type=int, default=0)
    args = parser.parse_args(argv)
    if args.command == "schedule":
        schedule_daily(args.workspace, args.hour, args.minute)
        return 0
    state = CollectionAgent(args.workspace).run(mode="manual", resume_run_id=args.resume_run_id)
    print(json.dumps(asdict(state), indent=2))
    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
